"""Export service for multi-format job data reports."""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReportBuilder:
    """Build multi-sheet Excel reports with color-coded scoring and summaries."""

    # Score band color mapping
    COLOR_MAP = {
        "Strong Match": "90EE90",      # Light green for ≥85
        "Good Match": "FFFFE0",         # Light yellow for ≥70
        "Fair Match": "FFE4B5",         # Moccasin for ≥50
        "Weak Match": "FFB6C1",         # Light pink for ≥35
        "Poor Match": "FFA07A",         # Light salmon for <35
    }

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def build_excel(
        self,
        jobs_df: pd.DataFrame,
        filtered_jobs_df: Optional[pd.DataFrame] = None,
        summary_stats: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        Build a multi-sheet Excel report.

        Args:
            jobs_df: DataFrame with all job matches
            filtered_jobs_df: DataFrame with filtered/rejected jobs (optional)
            summary_stats: Pre-computed summary stats (optional; will auto-compute if not provided)

        Returns:
            Excel file bytes ready for download
        """
        # Sheet 1: Job Matches with color-coded scores
        self._add_matches_sheet(jobs_df)

        # Sheet 2: Summary & Stats
        if summary_stats is None:
            summary_stats = self._compute_summary_stats(jobs_df, filtered_jobs_df)
        self._add_summary_sheet(summary_stats)

        # Sheet 3: Filtered Out (if provided)
        if filtered_jobs_df is not None and not filtered_jobs_df.empty:
            self._add_filtered_sheet(filtered_jobs_df)

        # Convert workbook to bytes
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _add_matches_sheet(self, df: pd.DataFrame) -> None:
        """Add Job Matches sheet with color-coded rows by fit_band."""
        ws = self.wb.create_sheet("Job Matches")

        # Select export columns
        export_cols = [
            "company",
            "title",
            "location",
            "score",
            "fit_band",
            "source",
            "source_lane_label",
            "salary_text",
            "work_type",
            "matched_keywords",
            "decision_reason",
            "url",
        ]
        # Filter to columns that exist
        available_cols = [c for c in export_cols if c in df.columns]
        export_df = df[available_cols].copy()

        # Write headers
        for col_idx, col_name in enumerate(available_cols, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name.replace("_", " ").title()
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write data rows with color-coding by fit_band
        for row_idx, (_, row) in enumerate(export_df.iterrows(), start=2):
            fit_band = row.get("fit_band", "")
            bg_color = self.COLOR_MAP.get(fit_band, "FFFFFF")

            for col_idx, col_name in enumerate(available_cols, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row[col_name]

                # Format URLs as links
                if col_name == "url" and pd.notna(value) and isinstance(value, str):
                    cell.hyperlink = value
                    cell.value = "Open Job"
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.value = value

                # Apply background color
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Auto-fit column widths
        self._auto_fit_columns(ws)

    def _add_summary_sheet(self, stats: Dict[str, Any]) -> None:
        """Add Summary sheet with fit band counts and top companies."""
        ws = self.wb.create_sheet("Summary")

        row = 1

        # Fit band summary
        ws.cell(row=row, column=1).value = "Fit Band Summary"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        band_counts = stats.get("fit_band_counts", {})
        for band in ["Strong Match", "Good Match", "Fair Match", "Weak Match", "Poor Match"]:
            count = band_counts.get(band, 0)
            ws.cell(row=row, column=1).value = band
            ws.cell(row=row, column=2).value = count
            bg_color = self.COLOR_MAP.get(band, "FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            ws.cell(row=row, column=2).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            row += 1

        row += 1

        # Source lane summary
        ws.cell(row=row, column=1).value = "Source Lane Distribution"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        source_counts = stats.get("source_lane_counts", {})
        for lane, count in sorted(source_counts.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1).value = lane
            ws.cell(row=row, column=2).value = count
            row += 1

        row += 1

        # Top companies
        ws.cell(row=row, column=1).value = "Top Companies"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        top_companies = stats.get("top_companies", [])
        for company, count in top_companies[:10]:
            ws.cell(row=row, column=1).value = company
            ws.cell(row=row, column=2).value = count
            row += 1

        # Auto-fit columns
        self._auto_fit_columns(ws)

    def _add_filtered_sheet(self, df: pd.DataFrame) -> None:
        """Add Filtered Out sheet with rejected jobs and reasons."""
        ws = self.wb.create_sheet("Filtered Out")

        export_cols = [
            "company",
            "title",
            "location",
            "score",
            "decision_reason",
            "matched_keywords",
            "url",
        ]
        available_cols = [c for c in export_cols if c in df.columns]
        export_df = df[available_cols].copy()

        # Write headers
        for col_idx, col_name in enumerate(available_cols, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name.replace("_", " ").title()
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write data rows
        for row_idx, (_, row) in enumerate(export_df.iterrows(), start=2):
            for col_idx, col_name in enumerate(available_cols, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row[col_name]

                if col_name == "url" and pd.notna(value) and isinstance(value, str):
                    cell.hyperlink = value
                    cell.value = "Open Job"
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.value = value

                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Auto-fit columns
        self._auto_fit_columns(ws)

    @staticmethod
    def _auto_fit_columns(ws) -> None:
        """Auto-fit all columns in a worksheet."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _compute_summary_stats(
        jobs_df: pd.DataFrame,
        filtered_df: Optional[pd.DataFrame] = None,
    ) -> Dict[str, Any]:
        """Compute summary statistics from job dataframes."""
        stats = {}

        # Fit band counts
        fit_band_counts = jobs_df.get("fit_band", pd.Series()).value_counts().to_dict()
        stats["fit_band_counts"] = {
            "Strong Match": fit_band_counts.get("Strong Match", 0),
            "Good Match": fit_band_counts.get("Good Match", 0),
            "Fair Match": fit_band_counts.get("Fair Match", 0),
            "Weak Match": fit_band_counts.get("Weak Match", 0),
            "Poor Match": fit_band_counts.get("Poor Match", 0),
        }

        # Source lane distribution
        if "source_lane_label" in jobs_df.columns:
            stats["source_lane_counts"] = jobs_df["source_lane_label"].value_counts().to_dict()
        else:
            stats["source_lane_counts"] = {}

        # Top companies by job count
        if "company" in jobs_df.columns:
            top_companies = jobs_df["company"].value_counts().head(10).to_list()
            stats["top_companies"] = [(company, count) for company, count in jobs_df["company"].value_counts().items()][:10]
        else:
            stats["top_companies"] = []

        return stats
